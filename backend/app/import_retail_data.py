"""Memory-safe importer for UCI Online Retail data.

Defaults to 50,000 rows for Render's 256 MB instance. Set RETAIL_IMPORT_LIMIT=0
only on an instance with at least 512 MB RAM to import all available rows.
"""
import os
from decimal import Decimal
from itertools import groupby
from pathlib import Path
from tempfile import NamedTemporaryFile

import httpx
from openpyxl import load_workbook
from sqlalchemy import select, text

from .db import Base, SessionLocal, engine
from .models import Customer, Order, OrderItem, Product
from .schema_migrations import recalculate_product_sales_trends

UCI_URL = "https://archive.ics.uci.edu/ml/machine-learning-databases/00352/Online%20Retail.xlsx"
BATCH_INVOICES = 200


def import_limit() -> int:
    return max(0, int(os.getenv("RETAIL_IMPORT_LIMIT", "50000")))


def download_workbook() -> Path:
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        with httpx.stream("GET", UCI_URL, timeout=120) as response:
            response.raise_for_status()
            for chunk in response.iter_bytes():
                temp_file.write(chunk)
        return Path(temp_file.name)


def normalized_row(values: tuple) -> dict | None:
    invoice, sku, description, quantity, invoice_date, price, customer_id, country = values[:8]
    if customer_id is None or description is None or quantity is None or price is None:
        return None
    return {
        "invoice": str(invoice),
        "sku": str(sku),
        "description": str(description)[:300],
        "quantity": int(quantity),
        "date": invoice_date,
        "price": Decimal(str(price)),
        "customer": str(int(customer_id)),
        "country": str(country or "Unknown"),
    }


def update_operational_metrics(session) -> None:
    # These deterministic SQL calculations avoid holding the whole dataset in memory.
    session.execute(text("""
        WITH stats AS (
            SELECT customer_id, SUM(total) AS lifetime_value,
                   COALESCE(MAX(order_date) FILTER (WHERE status = 'completed'), MAX(order_date)) AS last_completed_at,
                   COUNT(*) FILTER (WHERE status = 'completed') AS completed_orders
            FROM orders GROUP BY customer_id
        ), ranked AS (
            SELECT customer_id, lifetime_value, last_completed_at,
                   PERCENT_RANK() OVER (ORDER BY last_completed_at DESC) AS recency_percentile,
                   PERCENT_RANK() OVER (ORDER BY completed_orders DESC) AS frequency_percentile
            FROM stats
        )
        UPDATE customers c SET
            lifetime_value = GREATEST(0, ranked.lifetime_value),
            last_purchase_at = ranked.last_completed_at,
            churn_risk = ROUND(LEAST(0.95, GREATEST(0.05,
                0.75 * ranked.recency_percentile + 0.25 * ranked.frequency_percentile
            ))::numeric, 2)
        FROM ranked WHERE c.id = ranked.customer_id
    """))
    session.execute(text("""
        UPDATE customers SET segment = CASE
            WHEN last_purchase_at IS NULL THEN 'insufficient_history'
            WHEN churn_risk >= 0.65 AND lifetime_value >= 250 THEN 'at_risk_high_value'
            WHEN churn_risk >= 0.65 THEN 'at_risk_lower_value'
            WHEN lifetime_value >= 250 THEN 'high_value_active'
            ELSE 'regular_active'
        END
    """))
    session.execute(text("""
        WITH stats AS (
            SELECT oi.product_id,
                   AVG(CASE WHEN o.status = 'cancelled' THEN 1.0 ELSE 0.0 END) AS cancellation_rate
            FROM order_items oi JOIN orders o ON o.id = oi.order_id
            GROUP BY oi.product_id
        )
        UPDATE products p SET cancellation_rate = ROUND(stats.cancellation_rate::numeric, 3)
        FROM stats WHERE p.id = stats.product_id
    """))
    session.commit()


def import_data() -> None:
    Base.metadata.create_all(engine)
    limit = import_limit()
    source_file = download_workbook()
    imported_rows = 0
    imported_invoices = 0
    try:
        workbook = load_workbook(source_file, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = (normalized_row(values) for values in worksheet.iter_rows(min_row=2, values_only=True))
        valid_rows = (row for row in rows if row is not None)
        with SessionLocal() as session:
            customer_cache: dict[str, int] = {}
            product_cache: dict[str, int] = {}
            for invoice, grouped_rows in groupby(valid_rows, key=lambda row: row["invoice"]):
                invoice_rows = list(grouped_rows)
                if limit and imported_rows >= limit:
                    break
                # Keep an invoice atomic. Truncating its line items would create an
                # incomplete order that a later idempotent run would skip.
                if limit and imported_rows + len(invoice_rows) > limit:
                    break
                existing_order = session.scalar(select(Order).where(Order.invoice_number == invoice))
                if existing_order:
                    imported_rows += len(invoice_rows)
                    continue
                first = invoice_rows[0]
                customer_id = customer_cache.get(first["customer"])
                if customer_id is None:
                    customer = session.scalar(select(Customer).where(Customer.external_id == first["customer"]))
                    if not customer:
                        customer = Customer(external_id=first["customer"], country=first["country"], email=os.getenv("DEMO_RECIPIENT_EMAIL", "temp66642@gmail.com"))
                        session.add(customer)
                        session.flush()
                    customer_id = customer.id
                    customer_cache[first["customer"]] = customer_id
                total = sum((row["quantity"] * row["price"] for row in invoice_rows), Decimal("0"))
                order = Order(invoice_number=invoice, customer_id=customer_id, order_date=first["date"], status="cancelled" if invoice.startswith("C") else "completed", total=total)
                session.add(order)
                session.flush()
                for row in invoice_rows:
                    product_id = product_cache.get(row["sku"])
                    if product_id is None:
                        product = session.scalar(select(Product).where(Product.external_sku == row["sku"]))
                        if not product:
                            product = Product(external_sku=row["sku"], name=row["description"], unit_price=row["price"])
                            session.add(product)
                            session.flush()
                        product_id = product.id
                        product_cache[row["sku"]] = product_id
                    session.add(OrderItem(order_id=order.id, product_id=product_id, quantity=row["quantity"], unit_price=row["price"]))
                imported_rows += len(invoice_rows)
                imported_invoices += 1
                if imported_invoices % BATCH_INVOICES == 0:
                    session.commit()
                    session.expunge_all()
            session.commit()
            update_operational_metrics(session)
            recalculate_product_sales_trends(engine)
        workbook.close()
    finally:
        source_file.unlink(missing_ok=True)
    print(f"Import completed: {imported_rows} rows across {imported_invoices} new invoices")


if __name__ == "__main__":
    import_data()
